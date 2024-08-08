import openpyxl


def get_case_list(file_path):
    case_list = []
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        case_list.append([])
        for col in range(1, ws.max_column + 1):
            case_list[row - 2].append(ws.cell(row=row, column=col).value)
    # 关闭工作簿
    wb.close()

    for i in range(0, len(case_list)):
        case_list[i][1] = case_list[i][1].split('\n')
        for j in range(0, len(case_list[i][1])):
            if ')' in case_list[i][1][j]:
                case_list[i][1][j] = case_list[i][1][j].replace(')', '')
                case_list[i][1][j] = case_list[i][1][j].split('(')
    return case_list


if __name__ == '__main__':
    x=get_case_list('UI自动化测试案例.xlsx')
    print(x)
