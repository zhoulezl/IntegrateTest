import time

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def make_excel(data: dict, false_list: list):
    data = data
    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active
    # 创建一个字体对象，并设置加粗
    bold_font = Font(bold=True)

    # 创建一个边框对象
    m = 'medium'
    thin_border = Border(left=Side(style=m), right=Side(style=m), top=Side(style=m), bottom=Side(style=m))

    # 创建一个填充对象，并设置背景色
    fill_title = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    fill_title_in = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid')
    fill_error = PatternFill(start_color='DA9694', end_color='DA9694', fill_type='solid')
    fill_error_in = PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid')
    fill_pass = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
    fill_pass_in = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
    # 写入数据到单元格
    date = time.localtime()
    ws['A1'] = f'{date.tm_year}年{date.tm_mon}月{date.tm_mday}日接口自动化测试执行报告'
    ws['A2'] = '模块名称'

    ws['D2'] = '调用次数'
    ws['F2'] = '失败次数'
    ws['H2'] = '通过率'
    title_list = ['测试时间', '测试场景', '测试步骤', '接口地址', '请求方式', '状态码', '测试结果', '错误信息',
                  '传入参数']
    wrong_title_list = ['模块名称', '测试场景', '测试步骤', '接口地址', '请求方式', '状态码', '测试结果', '错误信息',
                        '传入参数']
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 40
    # 合并单元格
    ws.merge_cells('A1:I1')
    ws.merge_cells('A2:C2')
    ws.merge_cells('D2:E2')
    ws.merge_cells('F2:G2')
    ws.merge_cells('H2:I2')
    ws.cell(row=1, column=1).font = Font(size=18, bold=True)
    for col in range(1, 10):
        ws.cell(row=1, column=col).fill = fill_title
    for col in range(1, 10):
        ws.cell(row=2, column=col).fill = fill_title_in
    ws.row_dimensions[1].height = 60

    ws.cell(row=2, column=1).font = Font(size=14)
    ws.cell(row=2, column=4).font = Font(size=14)
    ws.cell(row=2, column=6).font = Font(size=14)
    ws.cell(row=2, column=8).font = Font(size=14)
    rol_num = 3
    rol_left = 0
    rol_out = []

    col_num_list_title = [1, 4, 6, 8]
    col_merge_list = [[1, 3], [4, 5], [6, 7], [8, 9]]
    # 写入总览表头
    for key in data.keys():
        if key != "总计":
            for col_num_i in range(1, len(col_num_list_title)):
                ws.cell(row=rol_num, column=col_num_list_title[col_num_i], value=data[key][0][col_num_i - 1])
                ws.cell(row=rol_num, column=col_num_list_title[col_num_i]).font = Font(size=14)
                ws.cell(row=rol_num, column=col_num_list_title[col_num_i]).fill = fill_title_in
                ws.row_dimensions[rol_num].height = 30
            for i in range(len(col_merge_list)):
                ws.merge_cells(start_row=rol_num, start_column=col_merge_list[i][0],
                               end_column=col_merge_list[i][1], end_row=rol_num)
            ws.cell(row=rol_num, column=1, value=key).font = Font(size=14)
            ws.cell(row=rol_num, column=1, value=key).fill = fill_title_in
            rol_num += 1
    # 写入总览表头总计
    ws.cell(row=rol_num, column=1, value='总计').font = Font(size=14)
    ws.cell(row=rol_num, column=1).fill = fill_title_in
    for col_num_i in range(1, len(col_num_list_title)):
        ws.cell(row=rol_num, column=col_num_list_title[col_num_i], value=data['总计'][col_num_i - 1])
        ws.cell(row=rol_num, column=col_num_list_title[col_num_i]).font = Font(size=14)
        ws.cell(row=rol_num, column=col_num_list_title[col_num_i]).fill = fill_title_in
        ws.row_dimensions[rol_num].height = 30
    for i in range(len(col_merge_list)):
        ws.merge_cells(start_row=rol_num, end_row=rol_num, start_column=col_merge_list[i][0],
                       end_column=col_merge_list[i][1])
    rol_num += 1

    # 开始写入错误数据
    ws.cell(row=rol_num, column=1, value='出错接口详情如下:').font = Font(size=14, bold=True)
    for col in range(1, 10):
        ws.cell(row=rol_num, column=col).fill = fill_error
    ws.row_dimensions[rol_num].height = 30
    ws.merge_cells(f'A{rol_num}:I{rol_num}')
    rol_out.append(rol_num)
    rol_num += 1
    rol_left = rol_num
    for col_num in range(1, 10):
        ws.cell(row=rol_num, column=col_num).value = wrong_title_list[col_num - 1]
        ws.cell(row=rol_num, column=col_num).font = Font(size=14)
        if col_num == 7:
            ws.cell(row=rol_num, column=col_num).fill = fill_error_in
    rol_out.append(rol_num)
    rol_num += 1
    for false_data in false_list:

        for col_num in range(1, 10):
            ws.cell(row=rol_num, column=col_num, value=(false_data[col_num - 1]))
            if col_num == 7:
                ws.cell(row=rol_num, column=col_num).fill = fill_error_in
        rol_num += 1

    # 开始写入数据
    data.pop('总计')
    for key in data.keys():
        ws.cell(row=rol_num, column=1).value = key
        ws.cell(row=rol_num, column=1).font = Font(size=14, bold=True)
        ws.row_dimensions[rol_num].height = 30
        for col in range(1, 10):
            ws.cell(row=rol_num, column=col).fill = fill_pass
        ws.merge_cells(f'A{rol_num}:I{rol_num}')
        rol_num += 1
        rol_out.append(rol_num)
        for col_num in range(1, 10):
            ws.cell(row=rol_num, column=col_num).value = title_list[col_num - 1]
            ws.cell(row=rol_num, column=col_num).font = Font(size=14)
            # ws.cell(row=rol_num, column=col_num).fill = fill_pass_in
        rol_num += 1
        for row in data[key][1]:
            for col_num in range(1, 10):
                ws.cell(row=rol_num, column=col_num, value=(row[col_num - 1]))
                # ws.cell(row=rol_num, column=col_num).fill = fill_pass_in
            rol_num += 1

    # 设置对齐方式（居中）
    for col_num in range(1, 10):
        for rol_num in range(1, wb.worksheets[0].max_row + 1):
            ws.cell(row=rol_num, column=col_num).border = thin_border
            ws.cell(row=rol_num, column=col_num).alignment = Alignment(horizontal='center', vertical='center')
            if col_num in (4, 8, 9) and rol_num >= rol_left and rol_num not in rol_out and ws.cell(row=rol_num,
                                                                                                   column=col_num).value != '-':
                ws.cell(row=rol_num, column=col_num).alignment = Alignment(horizontal='left', vertical='center',
                                                                           wrap_text=True)
    # 保存工作簿
    wb.save(fr'test_records\{date.tm_year}年{date.tm_mon}月{date.tm_mday}日接口自动化测试执行报告.xlsx')
