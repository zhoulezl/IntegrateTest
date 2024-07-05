import json
import os
import openpyxl
import requests
import re

'''
import json
 
string = '{"name": "Alice", "age": 25, "city": "New York"}'
dictionary = json.loads(string)
'''


class Case:
    """
        todo
            self.case_name = case[0]
            self.case_id = case[1]
            self.case_pre_apis = 测试的接口列表
            self.case_pre_api_reqways = 每个接口的请求方式
            self.case_pre_api_resargs = 每个接口的请求参数
            self.request_after_actions = 测后接口列表
            self.actions_operate_list = 测后接口请求方式
            self.after_action_args = 测后接口请求参数列表
            self.case_assert = 断言条件
            self.case_designer = 案例设计人
            self.case_design_time = 案例设计时间
            self.case_isdone = False    接口是否执行
            self.case_status = False    测试是否通过
    """

    def __init__(self, case):
        self.case_name = case[0]
        self.case_id = case[1]
        self.case_pre_apis = case[2]
        self.case_pre_api_reqways = case[3]
        self.case_pre_api_reqargs = case[4]
        self.pre_request_list = []
        self.pre_response_list = []

        # self.case_sql_inneed = case[5]
        # self.case_sql_connection = case[6]
        # self.case_sql_query = case[7]

        self.case_assert = case[5]
        self.case_after_apis = case[6]
        self.case_after_api_reqways = case[7]
        self.case_after_api_reqargs = case[8]
        self.after_request_list = []
        self.after_response_list = []

        self.case_designer = case[9]
        self.case_design_time = case[10]
        self.case_isdone = False
        self.case_status = False

    def __str__(self):
        return self.case_name

    # 该方法的目的是为了让接口，方式，参数列表一一对应，使用一个下标即可
    # 另外，要去除参数中的大括号来方便操作
    def make_request_list(self, case_pre_apis, case_pre_api_reqways, case_pre_api_reqargs):
        request_list = []
        if len(case_pre_apis) == len(case_pre_api_reqways) == len(case_pre_api_reqargs):
            # 将case的请求进行整理
            for i in range(len(case_pre_apis)):
                request_list.append([])
                request_list[i].append(case_pre_apis[i])
                # 这里是为了去掉前后{}
                request_list[i].append(case_pre_api_reqways[i])
                request_list[i].append(case_pre_api_reqargs[i][1:len(case_pre_api_reqargs[i]) - 1])
        else:
            print(self.case_name + "，案例接口-请求方式-请求参数长度不一致!")
        return request_list

    def make_args_beauty_full_response(self, request_list, response_list):
        args_format = r'(pre -\d+-[a-zA-Z]+) | (after -\d+-[a-zA-Z]+)'
        for i in range(len(request_list)):
            args_list = request_list[i][2].split(',')
            for arg_num in range(len(args_list)):
                arg = args_list[arg_num]
                arg_key = arg.split(':')[0]
                arg_value = arg.split(':')[1]
                if re.match(args_format, arg_value):
                    way_apiNum_resKey = arg_value.split('-')
                    if way_apiNum_resKey[0] == 'pre':
                        arg_value = case.pre_response_list[int(way_apiNum_resKey[1] - 1)].get(way_apiNum_resKey[2])
                        if request_list[i][1] == 'POST':
                            arg = arg_key + ':' + arg_value
                        elif request_list[i][1] == 'GET':
                            arg = arg_key + '=' + arg_value
                        args_list[arg_num] = arg

                    elif way_apiNum_resKey[0] == 'after':
                        arg_value = case.after_response_list[int(way_apiNum_resKey[1] - 1)].get(
                            way_apiNum_resKey[2])
                        if request_list[i][1] == 'POST':
                            arg = arg_key + ':' + arg_value
                        elif request_list[i][1] == 'GET':
                            arg = arg_key + '=' + arg_value
                        args_list[arg_num] = arg
                else:
                    if request_list[i][1] == 'POST':
                        arg = arg_key + ':' + arg_value
                    elif request_list[i][1] == 'GET':
                        arg = arg_key + '=' + arg_value
                    args_list[arg_num] = arg
                #     continue
            arg_json = ''
            if request_list[i][1] == "POST":
                arg_json = '{'
            elif request_list[i][1] == 'GET':
                arg_json = ''
            if request_list[i][1] == 'POST':
                for j in range(len(args_list)):
                    arg_json = arg_json + args_list[j] + ','
                arg_json = arg_json[:-1] + '}'
            elif request_list[i][1] == 'GET':
                for j in range(len(args_list)):
                    arg_json = arg_json + args_list[j] + '&'
                arg_json = arg_json[:-1]
            request_list[i][2] = arg_json
            response = do_request(request_list[i][0], request_list[i][1], request_list[i][2])
            print(self.case_name, '\n', [response.url, arg_json, response.status_code])
            response_pass = '请求失败，请核对IP端口参数信息'
            if response.status_code == 200:
                response_pass = '请求通过'
            response_list.append([response.url, response.status_code, response_pass])
            # 依次请求
            self.pre_response_list = response_list
        # return request_list
        return self.pre_response_list


def do_request(api, way, args):
    res_test_response = ''
    request_api = ''
    if way == 'POST':
        request_api = api
        payload = json.loads(args)
        res_test_response = requests.post(request_api, json=payload)
    elif way == 'GET':
        request_api = api + '?' + args
        res_test_response = requests.post(request_api)
    else:
        print("请求方式编写有误")
    # print(res_test_response)
    return res_test_response


# else:
#     print("案例接口-请求方式-请求参数长度不一致!")


# 1.读取全部xlsx文件列表
caseFlies = []
for root, dirs, files in os.walk('.', topdown=False):
    for name in files:
        str = os.path.join(root, name)
        if str.split('.')[-1] == 'xlsx':
            caseFlies.append(str.split('.\\')[1])

# 2.获取全部全部的用例
case_list = []
for caseFlie in caseFlies:
    workbook = openpyxl.load_workbook(caseFlie)
    sheets = workbook.sheetnames
    for sheet in sheets:
        for i_row in range(2, workbook[sheet].max_row + 1):
            case_list.append([])
            for i_col in range(1, workbook[sheet].max_column + 1):
                cell = workbook[sheet].cell(row=i_row, column=i_col)
                case_list[i_row - 2].append(cell.value)
                # print(cell.value)

# 3.获取全部case实例，并填充到case_obj_list中
case_obj_list = []
for case_str in case_list:
    # print(case_str)
    case = Case(case_str)
    case_obj_list.append(case)
    case.case_pre_apis = case.case_pre_apis.split("\n")
    case.case_pre_api_reqways = case.case_pre_api_reqways.split("\n")
    case.case_pre_api_reqargs = case.case_pre_api_reqargs.split("\n")
    case.case_after_apis = case.case_after_apis.split("\n")
    case.case_after_api_reqways = case.case_after_api_reqways.split("\n")
    case.case_after_api_reqargs = case.case_after_api_reqargs.split("\n")

# 开始测试阶段
for case in case_obj_list:
    # （1）.将所有的请求放入到pre_request列表中去
    case.pre_request_list = case.make_request_list(case.case_pre_apis, case.case_pre_api_reqways,
                                                   case.case_pre_api_reqargs)
    # （2）.修饰请求列表中的参数 （3）.执行case中的请求，并将结果按顺序放入到response列表中
    case.pre_request_list = case.make_args_beauty_full_response(case.pre_request_list, case.pre_response_list)
    # （4）.
    print(case.case_name, case.pre_response_list)
