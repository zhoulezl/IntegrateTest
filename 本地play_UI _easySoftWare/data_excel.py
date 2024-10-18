import time

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def make_excel(data: dict,big_key):
    data = data
    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active
    # 创建一个字体对象，并设置加粗
    bold_font = Font(bold=True)
    row_out = [1, 2]
    # 创建一个边框对象
    m = 'medium'
    thin_border = Border(left=Side(style=m), right=Side(style=m), top=Side(style=m), bottom=Side(style=m))

    # 创建一个填充对象，并设置背景色
    fill_title = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    fill_title_in = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid')
    fill_error = PatternFill(start_color='DA9694', end_color='DA9694', fill_type='solid')
    fill_error_in = PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid')
    fill_pass = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
    fill_pass_in = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
    # 写入数据到单元格
    date = time.localtime()
    ws['A1'] = f'{date.tm_year}年{date.tm_mon}月{date.tm_mday}日UI自动化测试执行报告'
    ws['A2'] = '模块名称'
    ws['C2'] = '场景数量'
    ws['D2'] = '失败场景数量'
    ws['E2'] = '通过率'
    title_list = ['测试时间', '用例编号', '测试场景', '测试步骤', '测试结果']
    wrong_title_list = ['模块名称', '用例编号', '测试场景', '测试步骤', '测试结果']
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 80

    # 合并单元格
    ws.merge_cells('A1:E1')
    ws.merge_cells('A2:B2')
    ws.cell(row=1, column=1).font = Font(size=18, bold=True)
    for col in range(1, 6):
        ws.cell(row=1, column=col).fill = fill_title
    for col in range(1, 6):
        ws.cell(row=2, column=col).fill = fill_title_in
    ws.row_dimensions[1].height = 60

    ws.cell(row=2, column=1).font = Font(size=14)
    ws.cell(row=2, column=3).font = Font(size=14)
    ws.cell(row=2, column=4).font = Font(size=14)
    ws.cell(row=2, column=5).font = Font(size=14)
    rol_num = 3
    # 处理数据
    # 写入总览表头
    all_false_count = 0
    all_false_data = {}
    for key in data.keys():
        false_count = 0
        if key != "总计":
            for case in data[key]:
                tmp = ''
                for things in case[3]:
                    tmp += things
                case[3] = tmp
                case[2] = case[2].replace(key + '-', '')
                if '测试失败' in case[-1]:
                    false_count += 1
                    if key in all_false_data.keys():
                        all_false_data[key].append(case)
                    else:
                        all_false_data[key] = [case]
            # 模块名称
            ws.cell(row=rol_num, column=1).font = Font(size=14)
            ws.cell(row=rol_num, column=1).value = key
            ws.merge_cells(start_row=rol_num, end_row=rol_num, start_column=1, end_column=2)
            # 场景数量
            ws.cell(row=rol_num, column=3).font = Font(size=14)
            ws.cell(row=rol_num, column=3).value = len(data[key])
            # 失败场景数量
            ws.cell(row=rol_num, column=4).font = Font(size=14)
            ws.cell(row=rol_num, column=4).value = false_count
            # 通过率
            ws.cell(row=rol_num, column=5).font = Font(size=14)
            if len(data[key]) == 0:
                ws.cell(row=rol_num, column=5).value = '-'
            else:
                ws.cell(row=rol_num, column=5).value = f'{((len(data[key]) - false_count) * 100 / len(data[key])):.2f}%'
            row_out.append(rol_num)
            rol_num += 1
        all_false_count += false_count
    # 写入总览表头总计
    # 模块名称
    ws.cell(row=rol_num, column=1).font = Font(size=14)
    ws.cell(row=rol_num, column=1).value = "总计"
    ws.merge_cells(start_row=rol_num, end_row=rol_num, start_column=1, end_column=2)
    # 场景数量
    ws.cell(row=rol_num, column=3).font = Font(size=14)
    ws.cell(row=rol_num, column=3).value = data["总计"][1]
    # 失败场景数量
    ws.cell(row=rol_num, column=4).font = Font(size=14)
    ws.cell(row=rol_num, column=4).value = all_false_count
    # 通过率
    ws.cell(row=rol_num, column=5).font = Font(size=14)
    ws.cell(row=rol_num, column=5).value = f'{((data["总计"][1] - all_false_count) * 100 / data["总计"][1]):.2f}%'
    row_out.append(rol_num)
    rol_num += 1
    # wb.save(fr'test_records/{date.tm_year}年{date.tm_mon}月{date.tm_mday}日UI自动化测试执行报告.xlsx')

    # 开始写入错误数据
    if all_false_count == 0:
        pass
    else:
        ws.cell(row=rol_num, column=1).font = Font(size=14)
        ws.cell(row=rol_num, column=1).value = '失败场景如下:'
        ws.cell(row=rol_num, column=1).fill = fill_error
        ws.merge_cells(start_row=rol_num, end_row=rol_num, start_column=1, end_column=5)
        row_out.append(rol_num)
        rol_num += 1
        for i in range(len(wrong_title_list)):
            ws.cell(row=rol_num, column=i + 1).font = Font(size=14)
            ws.cell(row=rol_num, column=i + 1).value = wrong_title_list[i]
        row_out.append(rol_num)
        rol_num += 1
        for key in all_false_data.keys():
            for i in range(len(all_false_data[key])):
                # 模块名称
                ws.cell(row=rol_num, column=1).font = Font(size=11)
                ws.cell(row=rol_num, column=1).value = key
                # 用例编号
                ws.cell(row=rol_num, column=2).font = Font(size=11)
                ws.cell(row=rol_num, column=2).value = all_false_data[key][i][1]
                # 测试场景
                ws.cell(row=rol_num, column=3).font = Font(size=11)
                ws.cell(row=rol_num, column=3).value = all_false_data[key][i][2]
                # 测试步骤
                ws.cell(row=rol_num, column=4).font = Font(size=11)
                ws.cell(row=rol_num, column=4).value = all_false_data[key][i][3]
                ws.row_dimensions[rol_num].height = 140
                # 测试结果
                ws.cell(row=rol_num, column=5).font = Font(size=11)
                ws.cell(row=rol_num, column=5).value = all_false_data[key][i][4]
                rol_num += 1
    data.pop('总计')
    # 开始写入数据
    for key in data.keys():
        ws.cell(row=rol_num, column=1).font = Font(size=14)
        ws.cell(row=rol_num, column=1).value = key
        ws.cell(row=rol_num, column=1).fill = fill_title
        ws.merge_cells(start_row=rol_num, end_row=rol_num, start_column=1, end_column=5)
        row_out.append(rol_num)
        rol_num += 1
        for i in range(len(title_list)):
            ws.cell(row=rol_num, column=i + 1).font = Font(size=11)
            ws.cell(row=rol_num, column=i + 1).value = title_list[i]
        row_out.append(rol_num)
        rol_num += 1
        for case in data[key]:
            for i in range(len(case)):
                # 模块名称
                ws.cell(row=rol_num, column=i + 1).font = Font(size=11)
                ws.cell(row=rol_num, column=i + 1).value = case[i]
                ws.row_dimensions[rol_num].height = 140
            rol_num += 1
            # 用例编号
            # 测试场景
            # 测试步骤
            # 测试结果

    # 设置对齐方式（居中）
    for col_num in range(1, 6):
        for rol_num in range(1, wb.worksheets[0].max_row + 1):
            ws.cell(row=rol_num, column=col_num).border = thin_border
            ws.cell(row=rol_num, column=col_num).alignment = Alignment(horizontal='center', vertical='center')
    for col_num in range(4, 6):
        for rol_num in range(1, wb.worksheets[0].max_row + 1):
            if rol_num not in row_out:
                ws.cell(row=rol_num, column=col_num).border = thin_border
                ws.cell(row=rol_num, column=col_num).alignment = Alignment(horizontal='left', vertical='center',
                                                                           wrapText=True)
    # 保存工作簿
    wb.save(fr'test_records/{big_key}-{date.tm_year}年{date.tm_mon}月{date.tm_mday}日UI自动化测试执行报告.xlsx')


if __name__ == '__main__':
    excel_data = {'modelers-数据集': [['测试用例001', '创建公有数据集创建文件下载并删除',
                                       ['登录sh环境\n', '鼠标悬停在个人头像上\n', '悬停方式创建数据集\n', '新建文件\n',
                                        '下载文件\n',
                                        '删除当前数据集\n', '退出登录\n'], '测试通过']],
                  'modelers-讨论区': [['测试用例002', '创建公有数据集创建文件下载并删除',
                                       ['登录sh环境\n', '鼠标悬停在个人头像上\n', '悬停方式创建数据集\n', '新建文件\n',
                                        '下载文件\n',
                                        '删除当前数据集\n', '退出登录\n'], '测试通过']],
                  '总计': ['总计', 2, 0, 100]}
    make_excel(excel_data)
